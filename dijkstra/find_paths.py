#!C:\Python27
# -*- coding: utf-8 -*-
import sys
import time
from openpyxl import load_workbook, Workbook
from modules import get_graph_builder, get_path_finder_strategy

"""
    This module is meant to be visible for the user and to be used directly
    with no interaction with other subpackages of freight_transport_network.

    Provides a single class (Network) and main methods that use it to find
    shortest paths between all nodes of a railway network and a roadway
    network, represented by a list of links and its distances (see "data"
    folder) for each gauge used in the network.

    Inputs ("data" folder):
        railway_links_table.xlsx: List of railway links and their distances,
            for each different gauge.
        roadway_links_table.xlsx: List of railway links and their distances,
            for a unique gauge (roadway mode doesn't usually have "gauges").

    Outputs ("paths" folder):
        railway_shortest_paths.xlsx: List of paths between all nodes of the
            network, by gauge.
        roadway_shortest_paths.xlsx: List of paths between all nodes of the
            network, for a unique gauge.
"""


class Network():

    """Represents a network with different gauges.

    A Network instance can have many different gauges, wich in practice is
    like having different isolated subnetworks. Each gauge of the network has
    its own Graph representing all the nodes of the network-gauge connected by
    its links."""

    PATH_SHEET_SUFFIX = "paths"
    PATH_FIELDS = ["id_od", "origin", "destination", "distance", "path",
                   "gauge"]

    def __init__(self):
        self.graphs = {}

    # PUBLIC
    def create_graphs(self, links):
        """Create graphs from lists of links.

        Links argument may be a Workbook with link tables or a dictionary with
        Link objects.

        Args:
            links (workbook): A workbook containing one worksheet per gauge or
                all gauges in a single worksheet.
            links (dictionary): A dictionary with gauges, and link ids as keys
                to access Link objects. links[gauge][id_link] = Link()
        """

        graph_builder = get_graph_builder(links)
        self.graphs = graph_builder.get_graphs(links)

    def find_shortest_paths(self, strategy_name, argument=None):
        """Find shortest paths for each possible pair of nodes, by gauge."""

        # start total networks timer
        total_timer_start = time.time()

        path_finder = get_path_finder_strategy(strategy_name)
        paths = path_finder.find_shortest_paths(self.graphs, argument)

        # stop total networks timer
        elapsed = (time.time() - total_timer_start)
        self._report_time(elapsed, "all gauges")

        return paths

    def find_shortest_path(self, id_od, strategy_name, argument=None):
        """Find shortest paths for each possible pair of nodes, by gauge."""

        path_finder = get_path_finder_strategy(strategy_name)
        node_a, node_b = self._id_od_to_nodes(id_od)
        paths = path_finder.find_shortest_path(node_a, node_b, self.graphs,
                                               argument)

        return paths

    def store_paths_in_excel(self, paths, xl_output=None):
        """Store found paths in excel."""

        print "\n Saving results in excel..."
        sys.stdout.flush()

        # create worksheet to store all results
        wb = Workbook(write_only=True)
        ws_all = wb.create_sheet()
        ws_all.title = "all_" + self.PATH_SHEET_SUFFIX

        # write field names
        ws_all.append(self.PATH_FIELDS)

        # store paths for each network
        for gauge in paths:
            gauge_paths = paths[gauge]
            self._store_network_paths(wb, gauge, gauge_paths, ws_all)

        # save workbook
        wb.save(xl_output or self.XL_OUTPUT)

        print "Finished."

    # PRIVATE
    # reporting methods
    def _store_network_paths(self, wb, gauge, paths, ws_all=None):
        """Copy paths to general worksheet and to gauge specific worksheet."""

        # create worksheet and write field names
        ws = wb.create_sheet()
        ws.title = gauge + "_" + self.PATH_SHEET_SUFFIX
        ws.append(self.PATH_FIELDS)

        # get nodes to iterate them
        nodes = sorted(paths.keys())

        # iterate throught paths, copying them to worksheet
        for node_a in nodes:
            for node_b in nodes:

                # get distance and path
                distance = paths[node_a][node_b]["distance"]
                list_path = paths[node_a][node_b]["path"]
                path = self._list_path_to_string(list_path)

                # create id of the origin destination pair
                id_od = self._nodes_to_id_od(node_a, node_b)

                # copy data to gauge worksheet
                data = [id_od, node_a, node_b, distance, path, gauge]
                ws.append(data)

                # copy data to global worksheet (with all gauges) if provided
                if ws_all:
                    ws_all.append(data)

    # auxiliar private methods
    def _report_time(self, time_spend, activity):
        print "{} took {:.2} seconds".format(activity, time_spend)

    def _list_path_to_string(self, list_path):
        """Convert a list of nodes representing a path in a string."""

        if list_path:
            list_path = [str(node).zfill(3) for node in list_path]
            string_path = "-".join(list_path)

        else:
            string_path = None

        return string_path

    def _nodes_to_id_od(self, node_a, node_b):
        return str(node_a) + "-" + str(node_b)

    def _id_od_to_nodes(self, id_od):
        return id_od.split("-")


def main(xl_input, xl_output, strategy_name="isolated_gauges", argument=None):
    """Find shortest paths between all nodes of a network, by gauge.

    Args:
        xl_input: List of links of a network, by gauge.
        gauge_names: Names of the different gauges that are in the network.
        xl_output: List of shortest paths between all nodes, by gauge.
    """

    # load list of links
    wb = load_workbook(xl_input)

    # create a Network object
    network = Network()

    # create graphs from links, find shortest paths and store then in excel
    network.create_graphs(wb)
    paths = network.find_shortest_paths(strategy_name, argument)
    network.store_paths_in_excel(paths, xl_output)

    # return network object, in case of the user wants to use it
    return network


def main_railway():
    """Find shortest paths for the railway network."""

    XL_INPUT = "data/railway_links_table.xlsx"
    XL_OUTPUT = "paths/railway_shortest_paths.xlsx"
    network = main(XL_INPUT, XL_OUTPUT)

    return network


def main_roadway():
    """Find shortest paths for the roadway network."""

    XL_INPUT = "data/roadway_links_table.xlsx"
    XL_OUTPUT = "paths/roadway_shortest_paths.xlsx"
    network = main(XL_INPUT, XL_OUTPUT)

    return network


if __name__ == "__main__":

    # parse arguments if called with arguments
    if len(sys.argv) == 4:
        xl_input = sys.argv[1]
        gauge_names = sys.argv[2].split("-")
        xl_output = sys.argv[3]

        main(xl_input, gauge_names, xl_output)

    # call methods using default arguments if none are passed
    else:
        main_railway()
        main_roadway()
