import unittest
from openpyxl import load_workbook
from graph import get_graph_builder
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
from modules.builder.components.link import RailwayLink


def assert_equal_without_order(expected_graph, graph):

    # first check if all edges of graph are in expected graph
    for node in graph:
        for edge in graph[node]:
            assert edge in expected_graph[node]

    # second check if all edges of expected graph are in graph
    for node in expected_graph:
        for edge in expected_graph[node]:
            assert edge in graph[node]


class GraphBuilderTestCase(unittest.TestCase):

    def setUp(self):
        self.expected_graph = {'a': [('b', 2), ('c', 3)],
                               'b': [('a', 2), ('d', 5), ('e', 2)],
                               'c': [('a', 3), ('e', 5)],
                               'd': [('b', 5), ('e', 1), ('z', 2)],
                               'e': [('b', 2), ('c', 5), ('d', 1), ('z', 4)],
                               'z': [('d', 2), ('e', 4)]}

    def test_SingleWorksheet(self):
        XL_LINKS = os.path.join(os.path.dirname(__file__),
                                "test/test_single_ws_no_gauges.xlsx")
        wb = load_workbook(XL_LINKS)
        ws = wb.active

        graph_builder = get_graph_builder(ws)
        graph = graph_builder.get_graphs(ws)

        self.assertEqual(self.expected_graph, graph)

    def test_SingleWorksheetWithGauges(self):
        XL_LINKS = os.path.join(os.path.dirname(__file__),
                                "test/test_single_ws_with_gauges.xlsx")
        wb = load_workbook(XL_LINKS)
        ws = wb.active

        graph_builder = get_graph_builder(ws)
        graphs = graph_builder.get_graphs(ws)
        graph = graphs["unique"]

        self.assertEqual(self.expected_graph, graph)

    def test_MultipleWorksheets(self):
        XL_LINKS = os.path.join(os.path.dirname(__file__),
                                "test/test_multiple_ws.xlsx")
        wb = load_workbook(XL_LINKS)

        graph_builder = get_graph_builder(wb)
        graphs = graph_builder.get_graphs(wb)

        graph = graphs["ancha"]
        self.assertEqual(self.expected_graph, graph)

        graph = graphs["media"]
        self.assertEqual(self.expected_graph, graph)

    def test_WorkbookSingleWorksheet(self):
        XL_LINKS = os.path.join(os.path.dirname(__file__),
                                "test/test_single_ws_no_gauges.xlsx")
        wb = load_workbook(XL_LINKS)

        graph_builder = get_graph_builder(wb)
        graph = graph_builder.get_graphs(wb)

        self.assertEqual(self.expected_graph, graph)

    def test_LinksDictionary(self):

        expected_graph = {'1': [('2', 2.0), ('3', 3.0)],
                          '2': [('1', 2.0), ('4', 5.0), ('5', 2.0)],
                          '3': [('1', 3.0), ('5', 5.0)],
                          '4': [('2', 5.0), ('5', 1.0), ('27', 2.0)],
                          '5': [('2', 2.0), ('3', 5.0), ('4', 1.0),
                                ('27', 4.0)],
                          '27': [('4', 2.0), ('5', 4.0)]}

        links = {}
        for node in expected_graph:
            for edge in expected_graph[node]:

                id = node + "-" + edge[0]
                links[id] = {}
                distance = edge[1]
                gauge = "ancha"
                link = RailwayLink(id, distance, gauge)
                links[id][gauge] = link

                gauge = "media"
                link = RailwayLink(id, distance, gauge)
                links[id][gauge] = link

        graph_builder = get_graph_builder(links)
        graphs = graph_builder.get_graphs(links)

        graph = graphs["ancha"]
        assert_equal_without_order(expected_graph, graph)

        graph = graphs["media"]
        assert_equal_without_order(expected_graph, graph)


if __name__ == '__main__':
    unittest.main()
