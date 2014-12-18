#!C:\Python27
# -*- coding: utf-8 -*-
from openpyxl import worksheet, Workbook


class Graph(dict):

    """Represents a graph (nodes with weighted links between them).

    It is built from a list of links taken from excel file. Is a subclass of
    a python dictionary where keys are all nodes of a network and values are
    lists of tuples (node_b, weight) representing the weighted links a
    particular node has.

    The representation of a network would be like this:

        {'a': [('b', 2), ('c', 3)],
         'b': [('a', 2), ('d', 5), ('e', 2)],
         'c': [('a', 3), ('e', 5)],
         'd': [('b', 5), ('e', 1), ('z', 2)],
         'e': [('b', 2), ('c', 5), ('d', 1), ('z', 4)],
         'z': [('d', 2), ('e', 4)]}
    """

    def add_edge(self, node_a, node_b, weight):
        """Add a node a to the graph with a weighted link with node b."""

        # create weighted link as a tuple
        weighted_edge = (node_b, float(weight))

        # create node_a if not already in graph
        if not node_a in self:
            self[node_a] = []

        # add link to node_a if not already present
        if not weighted_edge in self[node_a]:
            self[node_a].append(weighted_edge)


class BaseGraphBuilder(object):

    """docstring for BaseGraphBuilder"""

    @classmethod
    def _are_nodes_in_one_column(self, first_cell_value):
        if type(first_cell_value) == int:
            return False
        else:
            return "-" in first_cell_value

    def _split_nodes(self, node_a, node_b=None):

        if self._are_nodes_in_one_column(node_a):
            return (node_a.split("-")[0], node_a.split("-")[1])

        else:
            return (node_a, node_b)


class SingleWorksheet(BaseGraphBuilder):

    @classmethod
    def accepts(self, ws_links):

        has_gauges = None
        is_ws = (type(ws_links) == worksheet.worksheet.Worksheet)

        if is_ws:
            first_cell_value = ws_links.cell(column=1, row=2).value
            if self._are_nodes_in_one_column(first_cell_value):
                gauge_value = ws_links.cell(column=3, row=2).value
            else:
                gauge_value = ws_links.cell(column=4, row=2).value

            has_gauges = (gauge_value and gauge_value != "")

        return is_ws and not has_gauges

    def get_graphs(self, ws):
        """Build graph from an excel list of links."""

        graph = Graph()

        # check the columns where nodes are
        cell_node_a = ws.cell(column=1, row=2)
        if self._are_nodes_in_one_column(cell_node_a.value):
            cell_node_b = ws.cell(column=1, row=2)
            cell_weight = ws.cell(column=2, row=2)

        else:
            cell_node_b = ws.cell(column=2, row=2)
            cell_weight = ws.cell(column=3, row=2)

        # parse data of each link into the Graph
        while cell_node_a.value:

            # store data of the row
            node_a, node_b = self._split_nodes(cell_node_a.value,
                                               cell_node_b.value)
            weight = cell_weight.value

            # adds link in both ways (a to b, b to a)
            graph.add_edge(node_a, node_b, weight)
            graph.add_edge(node_b, node_a, weight)

            # move on to the next row
            cell_node_a = cell_node_a.offset(row=1, column=0)
            cell_node_b = cell_node_b.offset(row=1, column=0)
            cell_weight = cell_weight.offset(row=1, column=0)

        return graph


class SingleWorksheetWithGauges(BaseGraphBuilder):

    @classmethod
    def accepts(self, ws_links):

        has_gauges = None
        is_ws = (type(ws_links) == worksheet.worksheet.Worksheet)

        if is_ws:
            first_cell_value = ws_links.cell(column=1, row=2).value
            if self._are_nodes_in_one_column(first_cell_value):
                gauge_value = ws_links.cell(column=3, row=2).value
            else:
                gauge_value = ws_links.cell(column=4, row=2).value

            has_gauges = (gauge_value and gauge_value != "")

        return is_ws and has_gauges

    def get_graphs(self, ws):
        """Build graph from an excel list of links."""

        graphs = {}

        # check the columns where nodes are
        cell_node_a = ws.cell(column=1, row=2)
        if self._are_nodes_in_one_column(cell_node_a.value):
            cell_node_b = ws.cell(column=1, row=2)
            cell_weight = ws.cell(column=2, row=2)
            cell_gauge = ws.cell(column=3, row=2)

        else:
            cell_node_b = ws.cell(column=2, row=2)
            cell_weight = ws.cell(column=3, row=2)
            cell_gauge = ws.cell(column=4, row=2)

        # parse data of each link into the Graph
        while cell_node_a.value:

            # store data of the row
            node_a, node_b = self._split_nodes(cell_node_a.value,
                                               cell_node_b.value)
            weight = cell_weight.value
            gauge = cell_gauge.value

            graph = self._get_graph(graphs, gauge)

            # adds link in both ways (a to b, b to a)
            graph.add_edge(node_a, node_b, weight)
            graph.add_edge(node_b, node_a, weight)

            # move on to the next row
            cell_node_a = cell_node_a.offset(row=1, column=0)
            cell_node_b = cell_node_b.offset(row=1, column=0)
            cell_weight = cell_weight.offset(row=1, column=0)
            cell_gauge = cell_gauge.offset(row=1, column=0)

        return graphs

    def _get_graph(self, graphs, gauge):

        if not gauge in graphs:
            graphs[gauge] = Graph()

        return graphs[gauge]


class WorkbookSingleWorksheet(BaseGraphBuilder):

    """docstring for WorkbookWithGauges"""
    @classmethod
    def accepts(self, wb_links):
        return type(wb_links) == Workbook and len(wb_links.worksheets) == 1

    def get_graphs(self, wb_links):

        ws = wb_links.active
        graph_builder = get_graph_builder(ws)

        return graph_builder.get_graphs(ws)


class MultipleWorksheets(BaseGraphBuilder):

    @classmethod
    def accepts(self, wb_links):
        return type(wb_links) == Workbook

    def get_graphs(self, wb_links):

        graphs = {}

        for ws in wb_links:
            gauge = ws.title
            graphs[gauge] = SingleWorksheet().get_graphs(ws)

        return graphs


class LinksDictionary(BaseGraphBuilder):

    @classmethod
    def accepts(self, dict_links):

        is_dict = (type(dict_links) == dict)
        has_dict = (type(dict_links.values()[0]) == dict)

        return is_dict and has_dict

    def get_graphs(self, dict_links):

        graphs = {}

        for id_link in dict_links:
            for gauge in dict_links[id_link]:

                if not gauge in graphs:
                    graphs[gauge] = Graph()
                graph = graphs[gauge]

                link = dict_links[id_link][gauge]
                node_a = str(link.nodes[0])
                node_b = str(link.nodes[1])
                weight = link.dist

                graph.add_edge(node_a, node_b, weight)
                graph.add_edge(node_b, node_a, weight)

        return graphs


BUILDERS = [SingleWorksheet, SingleWorksheetWithGauges,
            WorkbookSingleWorksheet, MultipleWorksheets, LinksDictionary]


def get_graph_builder(links):

    RV = None

    for builder in BUILDERS:
        if builder.accepts(links):
            RV = builder()
            break

    return RV
