import unittest
import os
import find_paths
from openpyxl import load_workbook


def compare_cells(wb1, wb2):
    """Compare two excels based on row iteration."""

    # compare each cell of each worksheet
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        for row1, row2 in zip(ws1.iter_rows(), ws2.iter_rows()):
            for cell1, cell2 in zip(row1, row2):

                msg = str(cell1.value) + " != " + str(cell2.value)
                assert cell1.value == cell2.value, msg

    return True


@unittest.skip("integration skipped")
class FindPathsModuleIntegrationTestCase(unittest.TestCase):

    """Test finding shortest paths of railway mode."""

    def setUp(self):

        root_dir = os.path.dirname(__file__)

        # names of excels to be compared
        XL_LINKS = "test/test_railway_links_table.xlsx"
        XL_EXPECTED_PATHS = "test/test_railway_shortest_paths.xlsx"
        XL_PATHS = "test/railway_shortest_paths.xlsx"
        self.test_input_excel = os.path.join(root_dir, XL_LINKS)
        self.test_compare_excel = os.path.join(root_dir, XL_EXPECTED_PATHS)
        self.test_output_excel = os.path.join(root_dir, XL_PATHS)

        # run scraper method
        find_paths.main(self.test_input_excel, self.test_output_excel)

    def test_same_excels(self):

        wb1 = load_workbook(self.test_compare_excel, use_iterators=True)
        wb2 = load_workbook(self.test_output_excel, use_iterators=True)

        self.assertEqual(wb1.get_sheet_names(), wb2.get_sheet_names())
        self.assertTrue(compare_cells(wb1, wb2))


class IsolatedGaugesStrategyTestCase(unittest.TestCase):

    def setUp(self):

        G1 = {
            'a': [('b', 4), ('c', 2)],
            'b': [('a', 4), ('c', 1), ('d', 5)],
            'c': [('a', 2), ('b', 1), ('d', 8), ('e', 10)],
            'd': [('b', 5), ('c', 8), ('e', 2), ('z', 6)],
            'e': [('c', 10), ('d', 2), ('z', 3)],
            'z': [('d', 6), ('e', 3)],
        }

        self.network = find_paths.Network()
        self.network.graphs = {"unique": G1}

    def test_isolated_gauges(self):
        paths = self.network.find_shortest_paths("isolated_gauges")
        az_path = paths["unique"]["a"]["z"]["path"]
        self.assertEqual(az_path, ['a', 'c', 'b', 'd', 'e', 'z'])

    def test_restricted_nodes(self):
        paths = self.network.find_shortest_paths("isolated_gauges", ["e"])
        az_path = paths["unique"]["a"]["z"]["path"]
        self.assertEqual(az_path, ['a', 'c', 'b', 'd', 'z'])

    def test_restricted_links(self):
        paths = self.network.find_shortest_paths("isolated_gauges",
                                                 [("d", "e")])
        az_path = paths["unique"]["a"]["z"]["path"]
        self.assertEqual(az_path, ['a', 'c', 'b', 'd', 'z'])

    def test_find_shortest_path(self):
        paths = self.network.find_shortest_path("a-z", "isolated_gauges")
        az_path = paths["unique"]["path"]
        self.assertEqual(az_path, ['a', 'c', 'b', 'd', 'e', 'z'])


if __name__ == '__main__':
    unittest.main()
