from openpyxl import load_workbook
from railway_od import OD, Path
from railway_link import Link
from railway_parameters import Parameter


class BaseXlLoad():
    """Creates base class for iterate rows of a worksheet."""

    def __init__(self, xl_name):
        self.wb = load_workbook(xl_name, True)
        self.ws = self.wb.get_active_sheet()

    def __iter__(self):
        return self._iterate_rows()


class XlLoadOD(BaseXlLoad):
    """Creates an iterator of OD pairs from an excel workbook."""

    def _iterate_rows(self):

        # iterate trough rows creating and yielding od pairs
        for row in self.ws.iter_rows():

            # skip first row and empty rows
            if not row[0].row == 1 and row[0].value:

                # take field values
                id_od = row[0].value
                ton = row[1].value
                path = row[2].value
                gauge = row[3].value

                # create od pair
                od_pair = OD(id_od, ton, path, gauge)

                yield od_pair


class XlLoadLink(BaseXlLoad):
    """Creates an iterator of links from an excel workbook."""

    def _iterate_rows(self):

        # iterate trough rows creating and yielding od pairs
        for row in self.ws.iter_rows():

            # skip first row
            if not row[0].row == 1:

                # take field values
                id_link = row[0].value
                distance = row[1].value
                gauge = row[2].value

                # create link if all parameters are true
                if id_link and distance and gauge:
                    link = Link(id_link, distance, gauge)

                    yield link


class XlLoadParam(BaseXlLoad):
    """Creates an iterator of parameters from an excel workbook."""

    def _iterate_rows(self):

        # iterate sheets
        for ws in self.wb:

            # iterate trough rows creating and yielding od pairs
            for row in ws.iter_rows():

                # skip first row
                if not row[0].row == 1:

                    # take field values
                    id_param = row[0].value
                    value = row[1].value
                    desc = row[2].value

                    # create variable
                    parameter = Parameter(id_param, value, desc)

                    yield parameter


class XlLoadPath(BaseXlLoad):
    """Creates an iterator of paths from an excel workbook."""

    def _iterate_rows(self):

        # iterate trough rows creating and yielding od pairs
        for row in self.ws.iter_rows():

            # skip first row and skip empty id cells
            if row[0].value and not row[0].row == 1:

                # take field values
                id_path = row[0].value
                path = row[1].value
                gauge = row[2].value

                # create variable
                path = Path(id_path, path, gauge)

                yield path


def test():

    print "\nTest Case 1"
    print "-----------"
    railway_xl_input = "railway_od_pairs.xlsx"
    print railway_xl_input, "\n"
    i = 0
    for od_pair in XlLoadOD(railway_xl_input):
        print od_pair
        i += 1
        if i == 10:
            break

    print "______________________________________"

    print "\nTest Case 2"
    print "-----------"
    railway_xl_input = "railway_links.xlsx"
    print railway_xl_input, "\n"
    i = 0
    for link in XlLoadLink(railway_xl_input):
        print link
        i += 1
        if i == 10:
            break

    print "______________________________________"

    print "\nTest Case 3"
    print "-----------"
    railway_xl_input = "railway_parameters.xlsx"
    print railway_xl_input, "\n"
    i = 0
    for parameter in XlLoadParam(railway_xl_input):
        print parameter
        i += 1
        if i == 10:
            break

    print "______________________________________"

    print "\nTest Case 4"
    print "-----------"
    railway_xl_input = "railway_paths.xlsx"
    print railway_xl_input, "\n"
    i = 0
    for path in XlLoadPath(railway_xl_input):
        print path
        i += 1
        if i == 10:
            break

if __name__ == '__main__':
    test()
