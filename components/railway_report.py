from openpyxl import Workbook
from pprint import pprint


class Report():

    XL_REPORT = "railway_results_report.xlsx"

    def __init__(self, xl_report=None):
        """
        Args:
            xl_report: The path to excel file where results report will be
                stored.
        """
        self.xl_report = xl_report or self.XL_REPORT

    # PUBLIC
    def print_objects_report(self, rn):
        """Print report with examples of objects inside RailwayNetwork."""

        print "\nNumber of od_pairs:", len(rn.od_pairs)
        print "\nFirst 10 od_pairs"
        pprint(rn.od_pairs.values()[:10])

        print "\nFirst 10 od_pairs links"
        pprint([od_pair.links for od_pair in rn.od_pairs.values()[:10]])

        print "\nFirst 10 links ids"
        pprint(rn.links.keys()[:10])

        print "\nFirst 10 links values"
        pprint(rn.links.values()[:10])

        print "\nFirst 10 path values"
        pprint(rn.paths.values()[:10])

    def print_costs_report(self, rn):
        """Print only costs report of RailwayNetwork."""
        pass

    def report_to_excel(self, rn):
        """Make a report of RailwayNetwork results in excel."""

        # create a workbook
        wb = Workbook(write_only=True)

        # rolling material reports
        rn.locoms.report_to_excel(wb, "locoms")
        rn.wagons.report_to_excel(wb, "wagons")

        # network reports
        self._report_links_to_xl(rn, wb, "links")
        self._report_od_pairs_to_xl(rn, wb, "od_pairs")
        self._report_od_pairs_rej_to_xl(rn, wb, "od_pairs_rejected")
        self._report_global_results(rn, wb, "global_results")
        self._report_costs(rn, wb, "costs")

        # save excel report
        wb.save(self.xl_report)

    # PRIVATE
    def _report_global_results(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # copy function calls
        ws.append(["total tons", rn.get_total_tons()])
        ws.append(["total ton-km", rn.get_total_ton_km()])
        ws.append(["average distance",
                  rn.get_total_ton_km() / rn.get_total_tons()])
        ws.append(["rejected tons", rn.get_rejected_tons()])

    def _report_costs(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # add mobility costs to the costs sheet of report
        for key, value in rn.costs["mob"].items():
            ws.append([key, value])

        # add infrastructure costs to the costs sheet of report
        for key, value in rn.costs["inf"].items():
            ws.append([key, value])

    def _report_links_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        first_link = rn.links.values()[0].values()[0]
        ws.append(first_link.FIELDS)

        # iterate links appending data attributes of each link-gauge
        for link in rn.links.values():
            for gauge in link.values():
                link_attributes = gauge.get_attributes()
                ws.append(link_attributes)

    def _report_od_pairs_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        first_od = rn.od_pairs.values()[0]
        ws.append(first_od.FIELDS)

        # iterate od pairs appending data attributes of each od pair
        for od in rn.od_pairs.values():
            od_attributes = od.get_attributes()
            ws.append(od_attributes)

    def _report_od_pairs_rej_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        first_od = rn.od_pairs_rejected.values()[0]
        ws.append(first_od.FIELDS)

        # iterate rejected od pairs appending data attributes of each od pair
        for od in rn.od_pairs_rejected.values():
            od_attributes = od.get_attributes()
            ws.append(od_attributes)
