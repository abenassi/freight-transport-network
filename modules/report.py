from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, Alignment, Font
from pprint import pprint


class BaseReport():

    """Base report class to describe modal networks.

    It must be subclassed for RoadwayReport or RailwayReport to be used."""

    XL_GLOBAL_REPORT = "reports/freight_network_report.xlsx"
    WS_GLOBAL_NAME = "global_results"
    ALIGN = Alignment(vertical='center')
    ALIGN_HEADER = Alignment(horizontal='center', vertical='center',
                             wrap_text=True)
    CELL_STYLE = Style(alignment=ALIGN)
    FONT = Font(bold=True)
    HEADER_STYLE = Style(alignment=ALIGN_HEADER, font=FONT)

    def __init__(self, xl_report=None, description=None, append_report=True):
        """
        Args:
            xl_report: The path to excel file where results report will be
                stored.
        """
        self.xl_report = xl_report or self.XL_REPORT
        self.description = description
        self.append_report = append_report

    def print_objects_report(self, rn):
        """Print report with examples of objects inside RailwayNetwork."""

        print "\nNumber of od_pairs:", len(rn.od_pairs)
        print "\nFirst 10 od_pairs"
        pprint(list(rn.iter_od_pairs(10)))

        print "\nFirst 10 od_pairs links"
        pprint([od_pair.get_links() for od_pair in rn.iter_od_pairs(10)])

        print "\nFirst 10 links ids"
        pprint(rn.links.keys()[:10])

        print "\nFirst 10 links values"
        pprint(rn.links.values()[:10])

        print "\nFirst 10 path values"
        pprint(rn.paths.values()[:10])

    def print_global_results_report(self, rn):
        """Print report only with global results of network."""

        print "\n***Global results***\n"
        print "total tons", "{:,.2f}".format(rn.ton)
        print "total ton-km", "{:,.2f}".format(rn.ton_km)
        print "average distance", rn.ton_km / rn.ton

    def print_costs_report(self, rn):
        """Print only costs report of RailwayNetwork."""

        print "\n***Mobility cost results***\n"
        pprint(rn.costs["mob"])

        print "\n***infrastructure cost results***\n"
        pprint(rn.costs["inf"])

    def links_by_od_to_excel(self, paths, xl_links_by_od):
        """Write table of links by od pair to excel."""

        # create a workbook
        wb = Workbook(write_only=True)

        # create ws
        ws = wb.create_sheet()
        ws.title = "od_links"

        # write fields
        first_row = ["id_od", "id_link"]
        ws.append(first_row)

        # iterate paths
        for path in paths.values():

            # iterate links
            for link in path.get_links():

                row = [path.get_id(), link]
                ws.append(row)

        # save excel report
        wb.save(xl_links_by_od or self.XL_LINKS_BY_OD)

    @classmethod
    def create_new_global_report(self):
        """Create new workbook for global report."""

        wb = Workbook(write_only=True)
        wb.create_sheet(title=self.WS_GLOBAL_NAME)
        wb.save(self.XL_GLOBAL_REPORT)

    # PRIVATE
    def _report_links_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        first_link = rn.links.values()[0].values()[0]
        ws.append(first_link.FIELDS)

        # iterate links appending data attributes of each link-gauge
        for link in rn.iter_links():
            link_attributes = link.get_attributes()
            ws.append(link_attributes)

        # add auto filter
        col_letter = get_column_letter(ws.get_highest_column())
        ws.auto_filter.ref = "A1:" + col_letter + "1"

    def _report_od_pairs_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # check there are od pairs
        if rn.has_od_pairs():

            # write fields
            first_od = rn.od_pairs.values()[0].values()[0]
            ws.append(first_od.FIELDS)

            # iterate od pairs appending data attributes of each od pair
            for od in rn.iter_od_pairs():
                od_attributes = od.get_attributes()
                ws.append(od_attributes)

            # add auto filter
            col_letter = get_column_letter(ws.get_highest_column())
            ws.auto_filter.ref = "A1:" + col_letter + "1"

        else:
            ws.append(["There are no od pairs!"])

    def _report_global_results(self, rn, wb, ws_name):

        # append results to an existing report
        if self.append_report:

            # get ws
            ws = wb.get_sheet_by_name(ws_name)

            # get row and column to copy values
            i_col = ws.max_column + 1
            i_row = 2

            # write report description
            ws.cell(row=1, column=i_col).value = self.description

            # copy function calls
            values = [rn.ton,
                      rn.ton_km,
                      rn.average_distance,
                      rn.dimension,
                      rn.high_density_dimension,
                      rn.low_density_dimension,
                      rn.density]

            for value in values:
                ws.cell(row=i_row, column=i_col).value = value
                i_row += 1

        else:
            # create ws
            ws = wb.create_sheet()
            ws.title = ws_name
            ws.append(["variable", self.description])

            # copy function calls
            ws.append(["total tons", rn.ton])
            ws.append(["total ton-km", rn.ton_km])
            ws.append(["average distance", rn.average_distance])

            # copy dimensions
            ws.append(["total dimension", rn.dimension])
            ws.append(["high density dimension", rn.high_density_dimension])
            ws.append(["low density dimension", rn.low_density_dimension])
            ws.append(["density", rn.density])

    def _report_costs(self, rn, wb, ws_name):

        # append results to an existing report
        if self.append_report:

            # get ws
            ws = wb.get_sheet_by_name(ws_name)

            # get row and column to copy values
            i_col = ws.max_column + 1
            i_row = 2

            # write report description
            ws.cell(row=1, column=i_col).value = self.description

            # add each type of costs to the costs sheet of report
            for cost_type in rn.costs.keys():
                for key, value in rn.costs[cost_type].items():
                    ws.cell(row=i_row, column=i_col).value = value
                    i_row += 1

        else:
            # create ws
            ws = wb.create_sheet()
            ws.title = ws_name
            ws.append(["variable", self.description])

            # add each type of costs to the costs sheet of report
            for cost_type in rn.costs.keys():
                for key, value in rn.costs[cost_type].items():
                    ws.append([key, value])

    def _style_ws(self, ws):

        # style cells with values
        for col in ws.columns[1:]:
            for cell in col[1:]:
                cell.style = self.CELL_STYLE

        # styling first column
        for i_col in ws.column_dimensions:
            ws.column_dimensions[i_col].width = 15.0
        ws.column_dimensions["A"].width = 25.0

        # styling first row
        ws.row_dimensions[1].height = 60.0
        for cell in ws.rows[0]:
            cell.style = self.HEADER_STYLE

    def _get_wb_freight_network_report(self):
        """Return worksheet with freight network report."""

        # open or create the wb global report
        try:
            wb = load_workbook(self.XL_GLOBAL_REPORT)
        except:
            wb = Workbook(write_only=True)

        return wb

    def _make_freight_network_report(self, rn):
        """Make summary report into the global results report."""

        wb_fn = self._get_wb_freight_network_report()
        try:
            ws_fn = wb_fn.active
        except:
            ws_fn = wb_fn.create_sheet(title=self.WS_GLOBAL_NAME)

        # get row and column to copy values
        i_col = ws_fn.max_column + 1
        i_row = 3

        # write report description
        ws_fn.cell(row=1, column=i_col).value = self.description
        ws_fn.cell(row=2, column=i_col).value = rn.MODE_NAME

        # get values to be written
        values = [("Mobility", rn.costs["mob"]["total_mobility"]),
                  ("Infrastructure",
                   rn.costs["inf"]["total_infrastructure"]),
                  ("Time", rn.costs["time"]["total_time"]),
                  ("Total", rn.get_total_cost_tk()),
                  ("", ""),
                  ("Tons", rn.ton),
                  ("Ton-km", rn.ton_km),
                  ("Total modal cost", rn.get_total_cost()),
                  ("", ""),
                  ("", ""),
                  ("", ""),
                  ("Wagons per locomotive", rn.get_wagons_per_locomotive()),
                  ("Average distance", rn.average_distance),
                  ("Total network dimension", rn.dimension),
                  ("Average density", rn.density)]

        for value in values:

            # write field names if they are not written
            if i_col == 2:
                ws_fn.cell(row=i_row, column=1).value = value[0]

            ws_fn.cell(row=i_row, column=i_col).value = value[1]

            i_row += 1

        wb_fn.save(self.XL_GLOBAL_REPORT)


class RoadwayNetworkReport(BaseReport):

    XL_REPORT = "reports/roadway_report.xlsx"
    XL_LINKS_BY_OD = "reports/roadway_links_by_od.xlsx"

    # PUBLIC
    def report_to_excel(self, rn):
        """Make a report of RailwayNetwork results in excel."""

        self._make_roadway_network_report(rn)
        self._make_freight_network_report(rn)

    # PRIVATE
    def _make_roadway_network_report(self, rn):
        """Make the complete report of roadway network results in excel."""

        if self.append_report:
            try:
                # open the last report
                wb = load_workbook(self.xl_report)
            except:
                # create a new report
                wb = Workbook(write_only=True)
                self.append_report = False
        else:
            # create a new report
            wb = Workbook(write_only=True)

        # network reports
        self._report_global_results(rn, wb, "global_results")
        self._report_costs(rn, wb, "costs")
        self._report_links_to_xl(rn, wb, "links")
        self._report_od_pairs_to_xl(rn, wb, "od_pairs")

        # styling all the report if its in appending mode
        if self.append_report:
            for ws in wb:
                self._style_ws(ws)

        # save excel report
        wb.save(self.xl_report)


class RailwayNetworkReport(BaseReport):

    XL_REPORT = "reports/railway_report.xlsx"
    XL_LINKS_BY_OD = "reports/railway_links_by_od.xlsx"

    # PUBLIC
    def print_rolling_material_report(self, rn):
        """Print rolling material results."""

        print "\n***Wagons results***"
        print rn.wagons

        print "\n***Locomotives results***"
        print rn.locoms

    def report_to_excel(self, rn):
        """Make a report of RailwayNetwork results in excel."""

        self._make_railway_network_report(rn)
        self._make_freight_network_report(rn)

    # PRIVATE
    def _make_railway_network_report(self, rn):
        """Make the complete report of railway network results in excel."""

        if self.append_report:
            try:
                # open the last report
                wb = load_workbook(self.xl_report)
            except:
                # create a new report
                wb = Workbook(write_only=True)
                self.append_report = False
        else:
            # create a new report
            wb = Workbook(write_only=True)

        # rolling material reports
        rn.locoms.report_to_excel(wb, "locoms", self.description,
                                  self.append_report)
        rn.wagons.report_to_excel(wb, "wagons", self.description,
                                  self.append_report)

        # network reports
        self._report_global_results(rn, wb, "global_results")
        self._report_costs(rn, wb, "costs")
        self._report_links_to_xl(rn, wb, "links")
        self._report_od_pairs_to_xl(rn, wb, "od_pairs")

        # styling all the report if its in appending mode
        if self.append_report:
            for ws in wb:
                self._style_ws(ws)

        # save excel report
        wb.save(self.xl_report)
