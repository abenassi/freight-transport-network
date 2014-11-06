from openpyxl import Workbook
from pprint import pprint


class BaseReport():

    """Base report class to describe modal networks.

    It must be subclassed for RoadwayReport or RailwayReport to be used."""

    def __init__(self, xl_report=None):
        """
        Args:
            xl_report: The path to excel file where results report will be
                stored.
        """
        self.xl_report = xl_report or self.XL_REPORT

    def print_objects_report(self, rn):
        """Print report with examples of objects inside RailwayNetwork."""

        print "\nNumber of od_pairs:", len(rn.od_pairs)
        print "\nFirst 10 od_pairs"
        pprint(rn.od_pairs.values()[:10])

        print "\nFirst 10 od_pairs links"
        pprint([od_pair.links for od_pair in rn.od_pairs.values()[:10]])

        print "\nFirst 10 links ids"
        pprint(rn.links.keys()[:10])

        print "\nFirst 10 links values"
        pprint(rn.links.values()[:10])

        print "\nFirst 10 path values"
        pprint(rn.paths.values()[:10])

    def links_by_od_to_excel(self, paths, xl_links_by_od):
        """Write table of links by od pair to excel."""

        # create a workbook
        wb = Workbook(write_only=True)

        # create ws
        ws = wb.create_sheet()
        ws.title = "od_links"

        # write fields
        first_row = ["id_od", "id_link"]
        ws.append(first_row)

        # iterate paths
        for path in paths.values():

            # iterate links
            for link in path.get_links():

                row = [path.get_id(), link]
                ws.append(row)

        # save excel report
        wb.save(xl_links_by_od or self.XL_LINKS_BY_OD)

    # PRIVATE
    def _report_links_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        first_link = rn.links.values()[0].values()[0]
        ws.append(first_link.FIELDS)

        # iterate links appending data attributes of each link-gauge
        for link in rn.links.values():
            for gauge in link.values():
                link_attributes = gauge.get_attributes()
                ws.append(link_attributes)

    def _report_od_pairs_to_xl(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # write fields
        all_od_pairs = rn.od_pairs.values()
        first_od = all_od_pairs[0]
        ws.append(first_od.FIELDS)

        # iterate od pairs appending data attributes of each od pair
        for od in all_od_pairs:
            od_attributes = od.get_attributes()
            ws.append(od_attributes)

    def _report_global_results(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # copy function calls
        ws.append(["total tons", rn.get_total_tons()])
        ws.append(["total ton-km", rn.get_total_ton_km()])
        ws.append(["average distance",
                  rn.get_total_ton_km() / rn.get_total_tons()])

    def _report_costs(self, rn, wb, ws_name):

        # create ws
        ws = wb.create_sheet()
        ws.title = ws_name

        # add mobility costs to the costs sheet of report
        for key, value in rn.costs["mob"].items():
            ws.append([key, value])

        # add infrastructure costs to the costs sheet of report
        for key, value in rn.costs["inf"].items():
            ws.append([key, value])


class RoadwayReport(BaseReport):

    XL_REPORT = "reports/roadway_results_report.xlsx"
    XL_LINKS_BY_OD = "reports/roadway_links_by_od.xlsx"

    def report_to_excel(self, rn):
        """Make a report of RailwayNetwork results in excel."""

        # create a workbook
        wb = Workbook(write_only=True)

        # network reports
        self._report_links_to_xl(rn, wb, "links")
        self._report_od_pairs_to_xl(rn, wb, "od_pairs")
        self._report_global_results(rn, wb, "global_results")

        # save excel report
        wb.save(self.xl_report)
        

class RailwayReport(BaseReport):

    XL_REPORT = "reports/railway_results_report.xlsx"
    XL_LINKS_BY_OD = "reports/railway_links_by_od.xlsx"

    # PUBLIC
    def print_costs_report(self, rn):
        """Print only costs report of RailwayNetwork."""
        pass

    def report_to_excel(self, rn):
        """Make a report of RailwayNetwork results in excel."""

        # create a workbook
        wb = Workbook(write_only=True)

        # rolling material reports
        rn.locoms.report_to_excel(wb, "locoms")
        rn.wagons.report_to_excel(wb, "wagons")

        # network reports
        self._report_links_to_xl(rn, wb, "links")
        self._report_od_pairs_to_xl(rn, wb, "od_pairs")
        self._report_global_results(rn, wb, "global_results")
        self._report_costs(rn, wb, "costs")

        # save excel report
        wb.save(self.xl_report)
